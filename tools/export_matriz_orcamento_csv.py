#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Lê a folha MATRIZ de um orçamento/med BM (xlsx) e gera CSV compatível com
repo_cliente_itens_importar_csv: tipo, nome, codigo, unidade, valor_unitario, descricao.

Classificação (3 grupos pedidos):
  - produto: coluna "Coluna1" com MATERIAIS (capítulo 3.x, 5.x itens novos, etc.).
  - servico + descricao "Custo de funcionários / hora": unidade H (hora) ou texto com
    PLANTÃO / EQUIPE OPERACIONAL (custo operacional de equipe).
  - servico (demais): manutenção (1.x), gestão (2.x), projetos (4.x, ex. HR), etc.;
    itens 4.x recebem também "Projetos complementares" na descrição.

Uso:
  python tools/export_matriz_orcamento_csv.py "C:\\caminho\\arquivo.xlsx" > saida.csv
"""
from __future__ import annotations

import csv
import re
import sys
from decimal import Decimal


def _norm(s: object) -> str:
    if s is None:
        return ""
    t = str(s).strip()
    return t


def _is_number(x: object) -> bool:
    if x is None:
        return False
    if isinstance(x, (int, float)) and not isinstance(x, bool):
        return True
    if isinstance(x, Decimal):
        return True
    s = _norm(x)
    if s == "":
        return False
    try:
        float(s.replace(",", "."))
        return True
    except ValueError:
        return False


def _float_price(x: object) -> float:
    if isinstance(x, (int, float)):
        return float(x)
    s = _norm(x).replace(",", ".")
    return float(s)


def _item_key(item: object) -> str:
    if item is None:
        return ""
    return _norm(item)


def classify(item: object, col1: object, desc: object, un: object) -> tuple[str, str | None]:
    """Retorna (tipo, descricao_extra ou None). descricao_extra vai no campo descricao do CSV."""
    col1u = _norm(col1).upper()
    descu = _norm(desc).upper()
    unu = _norm(un).upper()
    ikey = _item_key(item)

    if "MATERIAIS" in col1u or (ikey.startswith("5.") and "MATERIAIS" in col1u):
        return "produto", None
    if ikey.startswith("5.") and "ITENS NOVOS" in col1u:
        return "produto", None

    # Custo de pessoal/hora: plantão e hora operacional (H). HR (ex.: projetos) vai em serviços gerais.
    labor = (
        unu == "H"
        or "PLANTÃO" in descu
        or "PLANTAO" in descu
        or "EQUIPE OPERACIONAL" in descu
    )
    if labor:
        return "servico", "Custo de funcionários / hora"

    if "MATERIAIS" not in col1u and (ikey.startswith("1.") or ikey.startswith("2.") or ikey.startswith("4.")):
        if ikey in ("1.0", "2.0", "4.0", "5.0") or (descu and descu == col1u and unu == ""):
            return "", None
        extra = "Projetos complementares" if ikey.startswith("4.") else None
        return "servico", extra

    if "MATERIAIS" in col1u or (ikey and ikey[0:1] == "3"):
        return "produto", None

    return "", None


def trim_fields(
    nome_full: str, desc_extra: str | None, col1: str, max_nome: int = 160, max_desc: int = 500
) -> tuple[str, str | None]:
    nome_full = nome_full.strip()
    nome = nome_full
    overflow = ""
    if len(nome) > max_nome:
        overflow = nome_full[max_nome - 1 :].strip()
        nome = nome_full[: max_nome - 1].rstrip() + "…"

    parts: list[str] = []
    if desc_extra:
        parts.append(desc_extra)
    if col1 and col1.strip():
        parts.append(f"Grupo orçamento: {col1.strip()}")
    if overflow:
        parts.append(overflow)
    desc = " | ".join(parts) if parts else ""
    if len(desc) > max_desc:
        desc = desc[: max_desc - 1].rstrip() + "…"
    return nome, desc if desc else None


def export_xlsx(path: str, out_writer: csv.writer) -> int:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "MATRIZ" not in wb.sheetnames:
        raise SystemExit(f"Folha MATRIZ não encontrada. Folhas: {wb.sheetnames}")
    ws = wb["MATRIZ"]
    out_writer.writerow(["tipo", "nome", "codigo", "unidade", "valor_unitario", "descricao"])
    n = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        item, col1, desc, un, preco = (list(row) + [None] * 5)[:5]
        desc_s = _norm(desc)
        col1_s = _norm(col1)
        un_s = _norm(un)
        if desc_s == "":
            continue
        if not _is_number(preco):
            continue
        if un_s == "":
            continue
        tipo, extra = classify(item, col1, desc, un)
        if not tipo:
            continue
        vu = _float_price(preco)
        cod = _item_key(item)
        nome_base = desc_s
        nome_f, desc_f = trim_fields(nome_base, extra, col1_s)
        out_writer.writerow(
            [
                tipo,
                nome_f,
                cod,
                un_s,
                f"{vu:.4f}".replace(".", ","),
                desc_f or "",
            ]
        )
        n += 1
    wb.close()
    return n


def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__.strip(), file=sys.stderr)
        raise SystemExit(2)
    path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else None
    if out_path:
        with open(out_path, "w", encoding="utf-8-sig", newline="") as fp:
            w = csv.writer(fp, delimiter=";", quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
            n = export_xlsx(path, w)
    else:
        sys.stdout.reconfigure(encoding="utf-8")
        w = csv.writer(sys.stdout, delimiter=";", quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
        n = export_xlsx(path, w)
    print(f"# {n} linhas gravadas", file=sys.stderr)


if __name__ == "__main__":
    main()
